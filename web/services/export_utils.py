"""
Export utilities for CSV and Excel formats
"""

import csv
import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def export_to_csv(credentials: List) -> str:
    """Export credentials to CSV format"""
    output = io.StringIO()
    
    fieldnames = [
        'Hostname',
        'Username',
        'Password',
        'Credential Type',
        'Account Type',
        'Resource Type',
        'Domain Name',
        'Source',
        'Last Updated'
    ]
    
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for cred in credentials:
        writer.writerow({
            'Hostname': cred.hostname or '',
            'Username': cred.username or '',
            'Password': cred.password or '',
            'Credential Type': cred.credential_type or '',
            'Account Type': cred.account_type or '',
            'Resource Type': cred.resource_type or '',
            'Domain Name': cred.domain_name or '',
            'Source': cred.source or 'SDDC_MANAGER',
            'Last Updated': cred.last_updated.strftime('%Y-%m-%d %H:%M:%S') if cred.last_updated else ''
        })
    
    return output.getvalue()


def export_to_excel(credentials: List, environment_name: str) -> bytes:
    """Export credentials to Excel format with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Hostname',
        'Username',
        'Password',
        'Credential Type',
        'Account Type',
        'Resource Type',
        'Domain Name',
        'Source',
        'Last Updated'
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_num, cred in enumerate(credentials, 2):
        ws.cell(row=row_num, column=1, value=cred.hostname or '')
        ws.cell(row=row_num, column=2, value=cred.username or '')
        ws.cell(row=row_num, column=3, value=cred.password or '')
        ws.cell(row=row_num, column=4, value=cred.credential_type or '')
        ws.cell(row=row_num, column=5, value=cred.account_type or '')
        ws.cell(row=row_num, column=6, value=cred.resource_type or '')
        ws.cell(row=row_num, column=7, value=cred.domain_name or '')
        ws.cell(row=row_num, column=8, value=cred.source or 'SDDC_MANAGER')
        ws.cell(row=row_num, column=9, value=cred.last_updated.strftime('%Y-%m-%d %H:%M:%S') if cred.last_updated else '')
    
    # Adjust column widths
    column_widths = [25, 30, 30, 18, 18, 18, 20, 18, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

